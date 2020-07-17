import pyautogui
import pyperclip
import time
import tkinter as tk
import sys
import openpyxl as xl
from tkinter import filedialog

from win32api import GetKeyState
from win32con import VK_CAPITAL, VK_NUMLOCK

if GetKeyState(VK_CAPITAL) == 1:
    pyautogui.press('capslock')
    print("Capslock turned off")
if GetKeyState(VK_NUMLOCK) == 1:
    pyautogui.press('numlock')
    print("Numlock turned off")


entries = []
def read_xl(filename):
    wb = xl.load_workbook(filename)
    sheet = wb.active
    print("Max Row = "+ str(sheet.max_row))
    print("Max Column = "+ str(sheet.max_column))
    for row in range(2, sheet.max_row + 1):
        dt = ""
        for column in range(1, sheet.max_column + 1):
            data = sheet.cell(row=row, column=column)
            cell = data.value
            if cell == "":
                cell = "N/A"
            dt = dt + "\t" + str(cell)
        entries.append(dt)
    print(len(entries))

in_val = input("Read Excel file??? [Y/N]")
if in_val == "Y" or in_val == "y":
    tk.Tk().withdraw()
    filename = filedialog.askopenfilename()
##    filename = "F:/pyproj/tesseract/Book1.xlsm"
    if filename == "":
        sys.exit()
    print("Reading excel")
    read_xl(filename)




    

def click_new():
    new_loc = None
    n = 0
    while new_loc == None:
        new_loc = pyautogui.locateOnScreen('attachments/new.png')
        time.sleep(0.4)
        n += 1
        if n == 20:
            print("Exiting. . . Image not found")
            root.destroy()
            sys.exit()
    loc = pyautogui.center(new_loc)
    x, y = loc
    pyautogui.click(x+180, y+40)


def click_id():
    id_loc = None
    n = 0
    while id_loc == None:
        id_loc = pyautogui.locateOnScreen('attachments/id.png')
        time.sleep(0.4)
        n += 1
        if n == 20:
            print("Exiting. . . Image not found")
            root.destroy()
            sys.exit()
    loc = pyautogui.center(id_loc)
    x, y = loc
    pyautogui.click(x+30, y+8)





root = tk.Tk()
def paste_key():
    time.sleep(0.3)
    click_new()
    time.sleep(0.3)
    click_id()
    copied_text = pyperclip.paste()
    copied_text = copied_text.split("\t")
    ##print(copied_text[17],copied_text[18])
    m = copied_text[17]
    n = copied_text[18]
    copied_text[17] = n
    copied_text[18] = m
    ##print(copied_text[17],copied_text[18])
    print(copied_text)
    pyautogui.PAUSE = 0
    for ct in copied_text:
        pyautogui.write(ct)
        pyautogui.press(['tab'])



def pp(event):
    p()
def p():
    time.sleep(0.5)
    pyautogui.hotkey('win', '1')
    time.sleep(0.1)
    paste_key()
    
def oo(event):
    o()
def o():
    ##RUN SHORTCUT ON EXCEL
    time.sleep(0.5)
    pyautogui.hotkey('win', '2')
    time.sleep(0.1)
    pyautogui.press('down')
    time.sleep(0.1)

    pyautogui.hotkey('ctrl', 'shift', 'right')

    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(0.1)
    pyautogui.hotkey('win', '1')
    time.sleep(0.1)
    paste_key()


def next_entry():
    s = e.get()
    global var1
    ## get position number of entry
    var1 = 0
    for ent in entries:
        if ent.find(s) > 0:
            break
        var1 += 1
    if var1 == len(entries):
        print("Not Found")
        e.delete(0, tk.END)
        e.insert(0, "Not Found")
    else:
        print(var1)
        ## RUN MACRO
        pyperclip.copy(entries[var1][1:])
        p()
        e.delete(0, tk.END)
        if var1+1 < len(entries):
            ent = entries[var1+1].split("\t")[1]
        else:
            ent = "COMPLETED"
        e.insert(0, ent)



def ii(event):
    i()
def i():
    next_entry()



B1 = tk.Button(root, text ="P. Paste", width=30, command = p)
B1.pack()
B2 = tk.Button(root, text ="O. XL Copy + Paste", width=30, command = o)
B2.pack()
root.bind('p', pp)
root.bind('P', pp)
root.bind('o', oo)
root.bind('O', oo)

if in_val == "Y" or in_val == "y":
    e = tk.Entry(root, width=25)
    e.pack()
    e.delete(0, tk.END)
    ent = entries[0].split("\t")[1]
    e.insert(0, ent)
    B3 = tk.Button(root, text ="I. Variable Paste", width=30, command = i)
    B3.pack()
    root.bind('i', ii)
    root.bind('I', ii)

root.mainloop()
